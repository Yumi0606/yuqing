from aip import AipNlp
import time
import os
import openpyxl
import threading
from datetime import datetime

# 添加情绪分析处理锁
emotion_lock = threading.Lock()

def add_emotion(file_path, analyze_only=False, timeout=6000):
    """为评论数据添加情感分析结果，添加了超时保护和详细日志输出
    
    参数:
        file_path: 文件路径
        api_key: 百度AI的API密钥
        analyze_only: 如果为True，只返回统计结果不修改文件
        timeout: API调用超时时间(秒)
        
    返回:
        result_stats: 包含统计信息的字典
    """
    # 创建结果统计字典
    result_stats = {
        "total_comments": 0,
        "processed_comments": 0,
        "positive_count": 0, 
        "negative_count": 0,
        "neutral_count": 0,
        "unknown_count": 0,
        "errors": [],
        "file_path": file_path,
        "success": False,
        "platform_stats": {},  # 平台统计
        "date_stats": {},      # 日期统计
        "needs_processing": False,  # 是否需要处理
        "api_error": False,
        "api_error_msg": ""
    }
    
    # 检查文件是否存在
    file_path = os.path.normpath(file_path)
    if not os.path.exists(file_path):
        result_stats["errors"].append(f"文件不存在: {file_path}")
        print(f"文件不存在: {file_path}")
        return result_stats
    
    try:
        # 初始化百度AI NLP客户端 - 提前创建以便在需要时立即使用
        APP_ID = "118369536"
        API_KEY = "tarNvDUxBb4K2MZsbU0JSwU6"
        SECRET_KEY = "D66H4GcJfikCMJPa3RZbRYhInwghytMM"
        client = AipNlp(APP_ID,  API_KEY, SECRET_KEY)
        
        # 使用 openpyxl 读取文件
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 获取标题行和必要的列索引
        headers = [cell.value for cell in ws[1]]
        columns = {
            'content': None,
            'platform': None,
            'emotion': None,
            'time': None
        }
        
        # 查找各列索引
        for idx, header in enumerate(headers, 1):
            if header in columns:
                columns[header] = idx
        
        # 检查是否有内容列
        if columns['content'] is None:
            result_stats["errors"].append("找不到评论内容列，文件不合法")
            print("找不到评论内容列，文件不合法")
            return result_stats
        
        # 获取总评论数
        total_rows = ws.max_row - 1  # 减去标题行
        result_stats["total_comments"] = total_rows
        
        # 检查是否需要添加情感列
        if columns['emotion'] is None and not analyze_only:
            result_stats["needs_processing"] = True
            columns['emotion'] = len(headers) + 1
            ws.cell(row=1, column=columns['emotion'], value="emotion")
            print(f"为文件 {file_path} 添加emotion列")
        
        # 增加检查：即使已有情感列，也需要检查是否有未知情感需要处理
        elif columns['emotion'] is not None and not analyze_only:
            # 检查是否有"未知"情感的评论需要处理
            unknown_count = 0
            for row in range(2, ws.max_row + 1):
                emotion = ws.cell(row=row, column=columns['emotion']).value
                if emotion is None or emotion == "" or emotion == "未知":
                    unknown_count += 1
                    
            if unknown_count > 0:
                result_stats["needs_processing"] = True
                print(f"文件 {file_path} 已有情感列，但有 {unknown_count} 条评论情感未知，需要处理")
        
        # 如果是只分析模式且已有情感列，直接统计不修改
        if analyze_only and columns['emotion'] is not None:
            # 统计已有情感数据
            for row in range(2, ws.max_row + 1):
                # 处理平台信息
                platform = "bilibili"  # 默认平台
                if columns['platform']:
                    platform_val = ws.cell(row=row, column=columns['platform']).value
                    if platform_val:
                        platform = platform_val
                
                if platform not in result_stats["platform_stats"]:
                    result_stats["platform_stats"][platform] = 0
                result_stats["platform_stats"][platform] += 1
                
                # 处理情感统计
                if columns['emotion']:
                    emotion = ws.cell(row=row, column=columns['emotion']).value or "未知"
                    if emotion == "积极":
                        result_stats["positive_count"] += 1
                    elif emotion == "消极":
                        result_stats["negative_count"] += 1
                    elif emotion == "中性":
                        result_stats["neutral_count"] += 1
                    else:
                        result_stats["unknown_count"] += 1
                
                # 处理日期统计
                _process_date_stats(result_stats, row, columns['time'], ws)
            
            result_stats["success"] = True
            result_stats["processed_comments"] = total_rows
            return result_stats
        
        # 创建API错误标志
        api_error = False
        api_error_msg = ""
        
        # 使用锁保护情感分析过程
        if result_stats["needs_processing"] and not analyze_only:
            with emotion_lock:
                print(f"开始处理文件: {file_path}，共 {total_rows} 条评论")
                processed = 0
                start_time = time.time()
                
                # 增加每次API调用的间隔时间，避免QPS限制
                api_call_delay = 0.5  # 每次调用间隔0.5秒，可根据情况调整
                
                # 处理每一行数据
                for row in range(2, ws.max_row + 1):
                    # 如果已经检测到API错误，立即退出循环
                    if api_error:
                        print(f"检测到严重API错误，立即停止处理，已完成{processed}/{total_rows}条评论")
                        break
                        
                    # 检查当前行是否已有明确情感
                    current_emotion = ws.cell(row=row, column=columns['emotion']).value
                    if current_emotion and current_emotion not in ["", "未知"]:
                        # 已有明确情感，统计但不再处理
                        if current_emotion == "积极":
                            result_stats["positive_count"] += 1
                        elif current_emotion == "消极":
                            result_stats["negative_count"] += 1
                        elif current_emotion == "中性":
                            result_stats["neutral_count"] += 1
                        continue
                    
                    # 检查是否已超时
                    if time.time() - start_time > timeout:
                        print(f"处理超时，已完成 {processed}/{total_rows} 条评论")
                        result_stats["errors"].append(f"情感分析超时，已处理 {processed} 条评论")
                        break
                    
                    # 获取评论内容
                    content = ws.cell(row=row, column=columns['content']).value
                    if not content:
                        continue
                    
                    # 进行情感分析
                    try:
                        # 添加延迟，避免触发API速率限制
                        time.sleep(api_call_delay)
                        
                        # 尝试进行API调用
                        result = client.sentimentClassify(content)
                        print(f"API返回: {result}")
                        
                        # 检查API错误
                        if "error_code" in result:
                            error_code = result["error_code"]
                            error_msg = result.get("error_msg", "未知错误")
                            
                            # 检查是否是严重的权限或配额错误
                            if error_code in [6, 17, 18, 19, 110, 111]:
                                api_error = True
                                api_error_msg = f"API严重错误: {error_code} - {error_msg}"
                                print(f"⚠️ 检测到严重API错误: {api_error_msg}")
                                print("立即停止所有API调用")
                                result_stats["errors"].append(api_error_msg)
                                # 保存当前进度后立即退出循环
                                try:
                                    wb.save(file_path)
                                    print(f"已在遇到错误前保存当前进度 ({processed}条)")
                                except Exception as save_e:
                                    print(f"保存进度失败: {str(save_e)}")
                                break
                            
                            # 非严重错误，继续处理但记录
                            print(f"遇到API错误: {error_code} - {error_msg}，跳过当前评论")
                            ws.cell(row=row, column=columns['emotion'], value="未知")
                            result_stats["unknown_count"] += 1
                            continue
                        
                        # 处理API结果
                        if "items" in result and result["items"]:
                            item = result["items"][0]
                            sentiment = item["sentiment"]
                            if sentiment == 0:
                                emotion = "消极"
                                result_stats["negative_count"] += 1
                            elif sentiment == 1:
                                emotion = "中性"
                                result_stats["neutral_count"] += 1
                            elif sentiment == 2:
                                emotion = "积极"
                                result_stats["positive_count"] += 1
                            else:
                                emotion = "未知"
                                result_stats["unknown_count"] += 1
                            
                            # 更新情感值
                            ws.cell(row=row, column=columns['emotion'], value=emotion)
                        else:
                            print(f"无法获取情感结果，API返回: {result}")
                            ws.cell(row=row, column=columns['emotion'], value="未知")
                            result_stats["unknown_count"] += 1
                    
                    except Exception as api_e:
                        error_msg = f"情感分析API错误: {str(api_e)}"
                        print(error_msg)
                        result_stats["errors"].append(error_msg)
                        ws.cell(row=row, column=columns['emotion'], value="未知")
                        result_stats["unknown_count"] += 1
                        
                        # 检查是否是严重错误
                        if any(indicator in str(api_e).lower() for indicator in 
                              ["error_code", "qps", "quota", "limit", "access", "permission"]):
                            api_error = True
                            api_error_msg = f"API严重错误: {str(api_e)}"
                            print(f"⚠️ 检测到可能的API限制错误: {api_error_msg}")
                            print("立即停止所有处理")
                            break
                    
                    processed += 1
                    # 每处理50条记录保存一次文件
                    if processed % 50 == 0:
                        try:
                            wb.save(file_path)
                            print(f"已处理 {processed}/{total_rows} 条评论，文件已保存")
                        except Exception as save_e:
                            print(f"保存文件失败: {str(save_e)}")
                
                # 保存处理结果
                try:
                    wb.save(file_path)
                    print(f"文件处理完成，共处理 {processed}/{total_rows} 条评论")
                    if not api_error:  # 只有在没有API错误时才标记为成功
                        result_stats["success"] = True
                except Exception as e:
                    print(f"保存文件失败: {str(e)}")
                    result_stats["errors"].append(f"保存文件失败: {str(e)}")
                
                result_stats["processed_comments"] = processed
    
    except Exception as e:
        error_msg = f"处理文件时出错: {str(e)}"
        print(error_msg)
        result_stats["errors"].append(error_msg)
        
        # 检查是否是API错误
        if any(indicator in str(e).lower() for indicator in 
              ["error_code", "qps", "quota", "limit", "access", "permission"]):
            api_error = True
            api_error_msg = str(e)
    
    # 添加API错误标志到结果中
    result_stats["api_error"] = api_error
    if api_error:
        result_stats["api_error_msg"] = api_error_msg
    
    return result_stats

# 辅助函数: 处理日期统计
def _process_date_stats(result_stats, row, time_idx, ws):
    if time_idx:
        time_val = ws.cell(row=row, column=time_idx).value
        if time_val:
            # 如果时间格式是datetime对象，转换为字符串
            if isinstance(time_val, datetime):
                time_val = time_val.strftime("%Y-%m-%d %H:%M:%S")
            try:
                date = time_val.split()[0]  # 只取日期部分
                if date not in result_stats["date_stats"]:
                    result_stats["date_stats"][date] = 0
                result_stats["date_stats"][date] += 1
            except (AttributeError, IndexError):
                # 忽略错误格式的日期
                pass

# 辅助函数: 更新情感统计
def update_emotion_stats(result_stats, emotion):
    if emotion == "积极":
        result_stats["positive_count"] += 1
    elif emotion == "消极":
        result_stats["negative_count"] += 1
    elif emotion == "中性":
        result_stats["neutral_count"] += 1
    else:
        result_stats["unknown_count"] += 1

# 修改读取文件函数，直接从spide_data目录读取
def load_exist_res_file(group_name=None):
    # 确保使用spide_data目录
    root_path = "./res_file/spide_data"
    if group_name:
        root_path = os.path.join(root_path, group_name)
    
    print(f"搜索路径: {root_path}")
    result_files = []
    for root, dirs, files in os.walk(root_path):
        for file in files:
            if file.endswith('.xlsx') and not '_video_info' in file:
                result_files.append(os.path.join(root, file))
    print(f"找到 {len(result_files)} 个文件")
    return result_files

def main():
    group_name = '游戏'
    # api_key = "hZuLZY3gotc0qscU4Rp5TFFcGlA0mN0z"
    # 多层遍历文件夹内的文件
    root_path = f"./res_file/spide_data/{group_name}"
    for root, dirs, files in os.walk(root_path):
        for file in files:
            if file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                if '_video_info' in file_path:
                    continue
                add_emotion(file_path)


if __name__ == "__main__":
    main()
