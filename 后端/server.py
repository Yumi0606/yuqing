from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from datetime import datetime
import openpyxl
import collections
from typing import List
import json
import os
import uvicorn
from fastapi.middleware.cors import CORSMiddleware
from datetime import datetime, timedelta
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
import time
import threading
from bilibili_api import search_bilibili_videos, get_video_comments
from data_handler import save_comments_to_excel, create_videos_folder
import random
from emotion_recognition import add_emotion
from queue import Queue
from contextlib import asynccontextmanager
import queue

# 添加 lifespan 上下文管理器
@asynccontextmanager
async def lifespan(app: FastAPI):
    # 启动爬取工作线程
    crawl_thread = threading.Thread(target=crawl_worker, daemon=True)
    crawl_thread.start()
    
    # 启动情感分析工作线程
    sentiment_thread = threading.Thread(target=sentiment_worker, daemon=True)
    sentiment_thread.start()
    
    # 为现有的方案安排任务
    keyword_groups = load_keyword_groups()
    for group in keyword_groups:
        schedule_crawling_task(group)
    
    yield  # 这里是应用运行期间
    
    # 清理资源
    print("应用正在关闭，执行清理...")

# 创建应用时传入 lifespan 参数 - 这一行是关键，确保应用使用我们的 lifespan
app = FastAPI(lifespan=lifespan)

origins = [
    "*",  # Allow all origins. Replace with specific domains for better security.
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class CreateKeywordGroup(BaseModel):
    group_name: str
    keywords: List[str]
    is_collecting: bool
    created_at: datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data_count: int = 0
    warning_count: int = 0
    immediate_collect: bool = False
    collection_frequency: int = 4  # 默认4小时
    collection_count: int = 100    # 默认爬取100条评论
    last_collected_at: str = None  # 上次爬取时间


class UpdateKeywordGroup(BaseModel):
    group_name: str
    new_group_info: dict


class DeleteKeywordGroup(BaseModel):
    group_name: str


class FilterSentimentData(BaseModel):
    group_name: str = None
    emotion_type: str = "全部"
    platform: str = "全部"
    start_time: str = None
    end_time: str = None


KEYWORD_GROUP_FILE = "./keyword_groups.json"
scheduler = BackgroundScheduler()
scheduler.start()

# 存储当前运行的任务ID
running_jobs = {}

# 创建线程锁和任务队列
crawl_lock = threading.Lock()
crawl_queue = Queue()
crawl_results = {}  # 存储爬取结果

# 添加情感分析任务队列和工作线程
sentiment_queue = queue.Queue()
sentiment_results = {}
sentiment_lock = threading.Lock()

def sentiment_worker():
    """后台情感分析工作线程，处理队列中的情感分析任务"""
    print("情感分析工作线程已启动，等待任务...")
    while True:
        try:
            print("等待情感分析队列中的新任务...")
            # 从队列获取任务
            task_id, file_path= sentiment_queue.get()
            print(f"收到情感分析任务 {task_id}: {file_path}，开始处理")
            
            # 使用锁保护情感分析过程
            with sentiment_lock:
                try:
                    result = add_emotion(file_path, analyze_only=False)
                    sentiment_results[task_id] = {
                        "status": "completed",
                        "result": result
                    }
                    print(f"情感分析任务 {task_id} 成功完成")
                except Exception as e:
                    print(f"情感分析任务 {task_id} 失败: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    sentiment_results[task_id] = {
                        "status": "failed",
                        "error": str(e)
                    }
            
            # 标记任务完成
            sentiment_queue.task_done()
            print(f"情感分析任务 {task_id} 处理完成")
        except Exception as e:
            print(f"情感分析工作线程出错: {str(e)}")
            import traceback
            traceback.print_exc()

# 添加爬取线程管理器
def crawl_worker():
    """后台爬取工作线程，处理队列中的爬取任务"""
    print("爬取工作线程已启动，等待任务...")
    while True:
        try:
            print("等待队列中的新任务...")
            # 从队列获取任务
            task_id, group_name, keywords, collection_count = crawl_queue.get()
            print(f"收到爬取任务 {task_id}: {group_name}，开始处理")
            
            # 使用锁保护爬取过程
            with crawl_lock:
                try:
                    result = crawl_data(group_name, keywords, collection_count)
                    crawl_results[task_id] = {
                        "status": "completed", 
                        "result": result
                    }
                    print(f"任务 {task_id} 成功完成")
                except Exception as e:
                    print(f"爬取任务 {task_id} 失败: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    crawl_results[task_id] = {
                        "status": "failed", 
                        "error": str(e)
                    }
            
            # 标记任务完成
            crawl_queue.task_done()
            print(f"爬取任务 {task_id} 处理完成")
        except Exception as e:
            print(f"爬取工作线程出错: {str(e)}")
            import traceback
            traceback.print_exc()

# 请求头，防止被识别为爬虫
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": "https://www.bilibili.com/",
    "Origin": "https://www.bilibili.com",
    "Cookie": "Cookie:buvid3=D4E92FD9-687C-A3AE-99B2-3141F559D64368029infoc; b_nut=1727599168; _uuid=31026AE78-88CC-2886-4CC6-8EFCD2795E9B68387infoc; enable_web_push=DISABLE; buvid4=BF2FF1DB-F153-7953-E8EA-F13205AE5F9268937-024092908-kGraga3wKv8RvNDiRPebUQ%3D%3D; rpdid=|(JJmYY|muuu0J'u~k~lmmu~R; header_theme_version=CLOSE; DedeUserID=1961954258; DedeUserID__ckMd5=574b68c6d0abdf66; CURRENT_QUALITY=80; fingerprint=77359313eb2f7eda225579542c5e4cb2; buvid_fp_plain=undefined; enable_feed_channel=DISABLE; buvid_fp=24b3a0fb39c79721bbd77f60d67b8dd3; CURRENT_FNVAL=2000; LIVE_BUVID=AUTO5817395209894496; PVID=1; bili_ticket=eyJhbGciOiJIUzI1NiIsImtpZCI6InMwMyIsInR5cCI6IkpXVCJ9.eyJleHAiOjE3NDA0NTk3NjEsImlhdCI6MTc0MDIwMDUwMSwicGx0IjotMX0.PMrfEguYpQOiXiQBM5OzB3PGl4TcEojt575dvReRXV0; bili_ticket_expires=1740459701; SESSDATA=a3f5cea0%2C1755752562%2C9452c%2A22CjDXiSYjpAs7q8DDSyeKmKzuL5Wpdb3Jt_opQ9I7C4EAr9PwWsqcfnLErDYttUqiit8SVnRmTk5sYnN0Qm1aYlJkZ2tLMmczOGRFNzBtMWxPZWFrM0R3aHNIc3BtU3FfbkM3VE9faWFUd1hMTFZ0NkxFUlR6WGJDWXowWWRyZWxwZXNkZ0lPYXRnIIEC; bili_jct=e9a835f0073f4dac2d4a325d99d74aaa; b_lsid=FF3FBE6C_1952CF9B199; bsource=search_bing; bp_t_offset_1961954258=1036710349685915648; home_feed_column=4; browser_resolution=400-747"
}


def load_keyword_groups():
    if os.path.exists(KEYWORD_GROUP_FILE):
        with open(KEYWORD_GROUP_FILE, "r") as file:
            try:
                return json.load(file)
            except Exception as e:
                print(f'Error loading keyword groups: {e}')
                return []
    return []


def save_keyword_groups(keyword_groups):
    with open(KEYWORD_GROUP_FILE, "w") as file:
        json.dump(keyword_groups, file, indent=4, default=str)


def crawl_data(group_name, keywords, collection_count):
    """
    根据关键词爬取B站数据
    
    :param group_name: 关键词组名称
    :param keywords: 关键词列表
    :param collection_count: 需要爬取的评论数量
    """
    print(f"开始爬取 {group_name} 的数据，关键词: {keywords}, 目标评论量: {collection_count}")
    
    # 确保每个视频最多爬取的评论数量不超过1000
    max_comments_per_video = 1000
    
    # 获取当前时间作为记录
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    total_comments_collected = 0
    warning_count = 0  # 记录预警次数
    emotion_stats = {
        "positive_total": 0,
        "negative_total": 0,
        "neutral_total": 0,
        "unknown_total": 0,
        "files_processed": 0,
        "emotion_errors": []
    }
    
    
    # 如果有多个关键词，平均分配要爬取的评论数量
    if len(keywords) > 0:
        comments_per_keyword = collection_count // len(keywords)
        # 确保至少每个关键词爬取1条评论
        comments_per_keyword = max(1, comments_per_keyword)
    else:
        # 如果没有关键词，直接返回
        print(f"关键词组 {group_name} 没有关键词，无法爬取")
        return
    
    # 为每个关键词爬取数据
    for idx, keyword in enumerate(keywords):
        # 计算当前关键词需要爬取的评论数
        # 最后一个关键词可能需要爬取更多以达到总目标
        if idx == len(keywords) - 1:
            remaining = collection_count - total_comments_collected
            keyword_target = max(remaining, comments_per_keyword)
        else:
            keyword_target = comments_per_keyword
            
        keyword_collected = 0  # 当前关键词已收集的评论数
        retry_count = 0  # 重试计数器
        max_retries = 3  # 最大重试次数
        
        try:
            print(f"为关键词 '{keyword}' 爬取约 {keyword_target} 条评论")
            
            # 创建存储目录
            output_dir = os.path.join(".", "res_file", "spide_data", group_name)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 搜索相关视频
            try:
                videos = search_bilibili_videos(HEADERS, keyword, num_results=10)
                if not videos:
                    print(f"关键词 '{keyword}' 未找到相关视频")
                    continue
            except Exception as e:
                print(f"搜索关键词 '{keyword}' 视频时出错: {e}")
                continue
            
            # 按评论数排序，优先爬取评论多的视频
            videos.sort(key=lambda x: int(x.get('comment_count', 0) or 0), reverse=True)
            
            # 遍历每个视频，爬取评论直到达到目标
            for video_idx, video in enumerate(videos):
                # 如果已经达到当前关键词的目标评论量，停止爬取
                if keyword_collected >= keyword_target:
                    break
                
                # 如果尝试了太多视频还是无法达到目标，也停止
                if video_idx >= 50:  # 限制每个关键词最多尝试50个视频
                    print(f"关键词 '{keyword}' 已尝试了50个视频，但未达到目标评论量")
                    break
                
                # 计算还需要爬取的评论数量
                comments_needed = min(max_comments_per_video, keyword_target - keyword_collected)
                
                # 如果这个视频的评论数太少，可能不值得爬取
                try:
                    comment_count = int(video.get('comment_count', 0) or 0)
                    if comment_count < 5:
                        print(f"视频 {video['bvid']} 评论数过少 ({comment_count})，跳过")
                        continue
                except (ValueError, TypeError):
                    # 如果评论数不是有效数字，设为0
                    comment_count = 0
                    
                try:
                    # 估算需要爬取的页数，每页约20条评论，但最少1页
                    max_pages = max(1, min(50, (comments_needed + 19) // 20))
                    
                    # 设置超时，避免无限等待
                    retry_attempts = 0
                    max_retry_attempts = 2
                    comments = []
                    
                    while retry_attempts < max_retry_attempts and not comments:
                        try:
                            print(f"爬取视频 {video['bvid']} 的评论，目标: {comments_needed}条")
                            comments = get_video_comments(HEADERS, video['aid'], video['bvid'], max_pages=max_pages)
                            
                            # 如果没有评论但应该有，可能是临时错误，重试
                            if not comments and comment_count > 0:
                                retry_attempts += 1
                                print(f"未找到评论，重试 ({retry_attempts}/{max_retry_attempts})...")
                                time.sleep(2)  # 等待一下再重试
                            elif not comments:
                                # 确实没有评论，跳过
                                print(f"视频 {video['bvid']} 没有评论")
                                break
                        except Exception as e:
                            retry_attempts += 1
                            print(f"爬取评论时出错: {e}，重试 ({retry_attempts}/{max_retry_attempts})...")
                            time.sleep(2)  # 等待一下再重试
                            
                    if not comments:
                        print(f"无法从视频 {video['bvid']} 获取评论，跳过")
                        continue
                    
                    # 限制评论数量
                    if len(comments) > comments_needed:
                        comments = comments[:comments_needed]
                    
                    # 保存评论数据
                    try:
                        # 直接保存评论数据到 spide_data 目录
                        file_path = save_comments_to_excel(comments, video, output_dir, group_name, keyword)
                        
                        # 验证文件是否存在
                        if not os.path.exists(file_path):
                            print(f"警告: 生成的文件不存在: {file_path}")
                            file_path = os.path.abspath(file_path) # 尝试获取绝对路径
                            print(f"尝试使用绝对路径: {file_path}")
                            if not os.path.exists(file_path):
                                print(f"错误: 无法找到文件，跳过情感分析")
                                continue
                        
                        # 进行情绪分析
                        try:
                            emotion_result = add_emotion(file_path)
                            if emotion_result["success"]:
                                print(f"情绪识别完成: {emotion_result['file_path']}")
                                print(f"情绪统计: 积极={emotion_result['positive_count']}, "
                                      f"消极={emotion_result['negative_count']}, "
                                      f"中性={emotion_result['neutral_count']}, "
                                      f"未知={emotion_result['unknown_count']}")
                                
                                # 更新总统计数据
                                emotion_stats["positive_total"] += emotion_result["positive_count"]
                                emotion_stats["negative_total"] += emotion_result["negative_count"]
                                emotion_stats["neutral_total"] += emotion_result["neutral_count"]
                                emotion_stats["unknown_total"] += emotion_result["unknown_count"]
                                emotion_stats["files_processed"] += 1
                                
                                # 检查是否需要生成预警
                                if emotion_result["negative_count"] > (emotion_result["processed_comments"] * 0.3):
                                    warning_count += 1
                                    print(f"检测到大量负面评论，生成预警")
                            else:
                                print(f"情绪识别失败: {emotion_result['errors']}")
                                emotion_stats["emotion_errors"].extend(emotion_result["errors"])
                        except Exception as e:
                            error_msg = f"情绪识别处理异常: {str(e)}"
                            print(error_msg)
                            emotion_stats["emotion_errors"].append(error_msg)
                        
                        # 更新已收集的评论数量
                        keyword_collected += len(comments)
                        total_comments_collected += len(comments)
                        print(f"已从视频 {video['bvid']} 爬取 {len(comments)} 条评论，"
                              f"当前关键词已爬取: {keyword_collected}/{keyword_target}，"
                              f"总计: {total_comments_collected}/{collection_count}")
                    except Exception as e:
                        print(f"保存评论数据时出错: {e}")
                        continue
                        
                except Exception as e:
                    print(f"处理视频 {video['bvid']} 时出错: {e}")
                    continue
        
        except Exception as e:
            print(f"爬取关键词 '{keyword}' 相关数据时出错: {e}")
            continue
    
    # 在完成爬取后，将情绪统计添加到日志
    print(f"情绪分析总结: 积极={emotion_stats['positive_total']}, "
          f"消极={emotion_stats['negative_total']}, "
          f"中性={emotion_stats['neutral_total']}, "
          f"未知={emotion_stats['unknown_total']}")
    
    if emotion_stats["emotion_errors"]:
        print(f"情绪分析中出现 {len(emotion_stats['emotion_errors'])} 个错误")
    
    # 更新关键词组的统计信息
    keyword_groups = load_keyword_groups()
    for group in keyword_groups:
        if group["group_name"] == group_name:
            group["last_collected_at"] = current_time
            group["data_count"] = group.get("data_count", 0) + total_comments_collected
            group["warning_count"] = group.get("warning_count", 0) + warning_count
            # 可以添加情绪统计数据到关键词组
            group["emotion_stats"] = group.get("emotion_stats", {})
            group["emotion_stats"]["positive"] = group["emotion_stats"].get("positive", 0) + emotion_stats["positive_total"]
            group["emotion_stats"]["negative"] = group["emotion_stats"].get("negative", 0) + emotion_stats["negative_total"]
            group["emotion_stats"]["neutral"] = group["emotion_stats"].get("neutral", 0) + emotion_stats["neutral_total"]
            break
    
    save_keyword_groups(keyword_groups)
    return {
        "comments_collected": total_comments_collected,
        "target_count": collection_count,
        "warnings_generated": warning_count,
        "emotion_stats": emotion_stats
    }


def schedule_crawling_task(group):
    """为关键词组安排爬取任务"""
    group_name = group["group_name"]
    
    # 如果已有任务，先移除
    if group_name in running_jobs:
        scheduler.remove_job(running_jobs[group_name])
        del running_jobs[group_name]
    
    if group.get("is_collecting", False):
        # 设置定时任务
        frequency_hours = group.get("collection_frequency", 4)
        job = scheduler.add_job(
            crawl_data,
            IntervalTrigger(hours=frequency_hours),
            args=[group_name, group["keywords"], group.get("collection_count", 100)],
            id=f"crawl_{group_name}"
        )
        running_jobs[group_name] = f"crawl_{group_name}"
        
        # 如果需要立即开始爬取
        if group.get("immediate_collect", False):
            # 使用线程执行爬取，避免阻塞API响应
            thread = threading.Thread(
                target=crawl_data,
                args=(group_name, group["keywords"], group.get("collection_count", 100))
            )
            thread.daemon = True
            thread.start()
            
            # 重置immediate_collect，避免重复爬取
            group["immediate_collect"] = False
            keyword_groups = load_keyword_groups()
            for kg in keyword_groups:
                if kg["group_name"] == group_name:
                    kg["immediate_collect"] = False
                    break
            save_keyword_groups(keyword_groups)


def load_exist_res_file(group_name):
    if group_name is None:
        root_path = f"./res_file/spide_data"
    else:
        root_path = f"./res_file/spide_data/{group_name}"
    res = list()
    for root, dirs, files in os.walk(root_path):
        for file in files:
            if file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                res.append(file_path)
    return res


def read_xlex(file_path):
    """
    读取 Excel 文件，转换为标准格式
    
    参数:
        file_path: Excel 文件路径
    返回:
        data: 字典列表，每个字典代表一行数据，键为列标题
            {"platform": "B站", "emotion": "积极", "time": "2023-06-15 10:30:45", "content": "这个视频太棒了！"}
        count: 数据行数
    """
    try:
        # 打开工作簿和工作表
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        
        # 获取表头
        headers = [cell.value for cell in ws[1]]
        
        # 准备结果容器
        data = []
        
        # 从第二行开始遍历（跳过表头）
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 创建当前行的字典
            row_dict = {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
            
            # 添加到结果列表
            data.append(row_dict)
        
        # 行数（不包括表头）
        count = len(data)
        
        # 关闭工作簿
        wb.close()
        
        # 返回结果
        return data, count
    
    except Exception as e:
        print(f"读取文件 {file_path} 时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # 返回空结果
        return [], 0


@app.post("/create_keyword_group/")
async def create_keyword_group(group: CreateKeywordGroup):
    keyword_groups = load_keyword_groups()
    if any(existing_group["group_name"] == group.group_name for existing_group in keyword_groups):
        raise HTTPException(status_code=400, detail="Group name already exists")
    
    group_dict = group.dict()
    keyword_groups.append(group_dict)
    save_keyword_groups(keyword_groups)
    
    # 安排爬取任务
    schedule_crawling_task(group_dict)
    
    return {"message": "Keyword group created successfully", "group": group, 'code': 200}


@app.post("/update_keyword_group/")
async def update_keyword_group(updated_group: dict):
    keyword_groups = load_keyword_groups()
    if updated_group['group_name'] not in [group["group_name"] for group in keyword_groups]:
        raise HTTPException(status_code=404, detail="Keyword group not found")
    
    for group in keyword_groups:
        if group["group_name"] == updated_group['group_name']:
            # 检查是否改变了收集状态或频率
            old_is_collecting = group.get("is_collecting", False)
            old_frequency = group.get("collection_frequency", 4)
            
            group.update(updated_group)
            
            # 如果收集状态或频率改变，需要重新安排爬取任务
            if (old_is_collecting != group.get("is_collecting", False) or 
                old_frequency != group.get("collection_frequency", 4)):
                schedule_crawling_task(group)
            
            break
    
    save_keyword_groups(keyword_groups)
    return {"message": "Keyword group updated successfully", "group": updated_group, 'code': 200}


@app.post("/delete_keyword_group/")
async def delete_keyword_group(group_name: DeleteKeywordGroup):
    group_name = group_name.group_name
    
    # 停止相关的爬取任务
    if group_name in running_jobs:
        scheduler.remove_job(running_jobs[group_name])
        del running_jobs[group_name]
    
    keyword_groups = load_keyword_groups()
    keyword_groups = [group for group in keyword_groups if group["group_name"] != group_name]
    save_keyword_groups(keyword_groups)
    return {"message": f"{group_name} deleted successfully", "code": 200}


@app.get("/keyword_groups/")
async def get_keyword_groups(group_name: str = None):
    keyword_groups = load_keyword_groups()
    if group_name:
        keyword_groups = [
            group for group in keyword_groups if group_name.lower() in group["group_name"].lower()
        ]
        if not keyword_groups:
            raise HTTPException(status_code=404, detail="No matching keyword groups found")
    
    # 清零数据计数，从文件重新计算
    for group in keyword_groups:
        group["data_count"] = 0
        for file in load_exist_res_file(group["group_name"]):
            data, count = read_xlex(file)
            group["data_count"] += count
    
    total_data_count = sum(group["data_count"] for group in keyword_groups)
    total_warning_count = sum(group.get("warning_count", 0) for group in keyword_groups)
    return {"keyword_groups": keyword_groups, 'code': 200, "total_data_count": total_data_count, "total_warning_count": total_warning_count}


# 手动触发爬取
@app.post("/trigger_collection/")
async def trigger_collection(group_name: str, collection_count: int = None):
    keyword_groups = load_keyword_groups()
    target_group = None
    
    for group in keyword_groups:
        if group["group_name"] == group_name:
            target_group = group
            break
    
    if not target_group:
        raise HTTPException(status_code=404, detail="Keyword group not found")
    
    # 如果前端传入了爬取数量，使用传入的值，否则使用关键词组中的默认值
    actual_count = collection_count if collection_count is not None else target_group.get("collection_count", 100)
    
    # 生成任务ID
    task_id = f"{group_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # 将任务加入队列
    crawl_queue.put((task_id, group_name, target_group["keywords"], actual_count))
    
    return {
        "message": f"Collection triggered for {group_name} with count {actual_count}", 
        "task_id": task_id,
        "code": 200
    }


# 舆情分析
@app.get('/sentiment_analysis/')
async def sentiment_analysis(group_name: str = None):
    file_list = load_exist_res_file(group_name)
    platform_res = collections.defaultdict(int)
    emotion_res = collections.defaultdict(int)
    date_res = collections.defaultdict(int)
    
    # print(f"开始加载舆情数据，文件数量: {len(file_list)}")
    
    # 处理所有文件
    for file in file_list:
        try:
            # print(f"读取文件: {file}")
            # 使用简化的 read_xlex 函数读取文件
            items, _ = read_xlex(file)
            for item in items:
                # 使用get方法安全地获取platform值，如果缺失则使用默认值
                platform = item.get('platform', '未知平台')
                platform_res[platform] += 1
                emotion_res[item.get('emotion', '未知')] += 1
                
                # 处理日期数据
                if 'time' in item and item['time']:
                    try:
                        date = item['time'].split()[0]  # 只取日期部分
                        date_res[date] += 1
                    except (AttributeError, IndexError):
                        # 如果time不是字符串或格式不对，跳过
                        pass
        except Exception as e:
            print(f"处理文件 {file} 时出错: {str(e)}")
            import traceback
            traceback.print_exc()
    
    # 统计结果处理
    total = sum(platform_res.values())
    if total > 0:
        platform_res_list = [{"name": key, "value": round(value / total * 100, 2)} for key, value in platform_res.items()]
    else:
        platform_res_list = []
    
    total = sum(emotion_res.values())
    emotion_res_count = [{"name": key, "value": value} for key, value in emotion_res.items()]
    
    if total > 0:
        emotion_res_list = [{"name": key, "value": round(value / total * 100, 2)} for key, value in emotion_res.items()]
    else:
        emotion_res_list = []
    
    today = datetime.now()
    seven_days_ago = today - timedelta(days=7)
    fourteen_days_ago = today - timedelta(days=14)
    thirty_days_ago = today - timedelta(days=30)

    # 为了安全起见，只处理格式正确的日期
    date_res_list = {
        "last_7_days": [],
        "last_14_days": [],
        "last_30_days": [],
    }
    
    for key, value in date_res.items():
        try:
            date_obj = datetime.strptime(key, "%Y-%m-%d")
            date_item = {"name": key, "value": value}
            
            if date_obj >= seven_days_ago:
                date_res_list["last_7_days"].append(date_item)
            if date_obj >= fourteen_days_ago:
                date_res_list["last_14_days"].append(date_item)
            if date_obj >= thirty_days_ago:
                date_res_list["last_30_days"].append(date_item)
        except ValueError:
            # 如果日期格式不正确，跳过
            continue
    
    print(f"舆情分析完成")
    print(emotion_res_count)
    return {
        'platform_res': platform_res_list, 
        'emotion_res': emotion_res_list, 
        'date_res': date_res_list, 
        'code': 200, 
        'emotion_res_count': emotion_res_count
    }


@app.post("/filter_sentiment_data/")
async def filter_sentiment_data(info: FilterSentimentData):
    group_name = info.group_name
    emotion_type = info.emotion_type
    platform = info.platform
    start_time = info.start_time
    end_time = info.end_time
    
    print(f"筛选请求参数: 组={group_name}, 情感={emotion_type}, 平台={platform}, 开始时间={start_time}, 结束时间={end_time}")
    
    file_list = load_exist_res_file(group_name)
    print(f"找到{len(file_list)}个文件")
    
    filtered_data = []
    total_data_count = 0
    positive_count = 0
    negative_count = 0
    today = datetime.now().strftime("%Y-%m-%d")
    today_new_count = 0

    # 平台名称映射表
    platform_aliases = {
        "bilibili": ["b站", "哔哩哔哩", "bilibili", "哔哩"],
        "douyin": ["抖音", "tiktok", "douyin"],
        "weibo": ["微博", "weibo"],
        "wechat": ["微信", "wechat"]
    }
    
    # 处理平台筛选
    platform_matches = set()
    if platform:  # 只有当platform不为空时才进行平台筛选
        platform_lower = platform.lower()
        for std_name, aliases in platform_aliases.items():
            if platform_lower == std_name.lower() or platform_lower in [a.lower() for a in aliases]:
                platform_matches.update([std_name.lower()] + [a.lower() for a in aliases])
                break
    
    # 处理每个文件
    for  file in file_list:
        try:
            items, count = read_xlex(file)
            
                
            for item in items:
                # 1. 时间筛选
                if "time" not in item or not item["time"]:
                    continue
                
                try:
                    item_date = item["time"].split()[0]
                    item_time = datetime.strptime(item_date, "%Y-%m-%d")
                    
                    # 应用时间过滤
                    if start_time and item_time < datetime.strptime(start_time, "%Y-%m-%d"):
                        continue
                    if end_time and item_time > datetime.strptime(end_time, "%Y-%m-%d"):
                        continue
                    
                    # 记录今日数据
                    if item_date == today:
                        today_new_count += 1
                except:
                    continue
                
                # 2. 平台筛选
                item_platform = item.get("platform", "未知").lower()
                if platform and platform_matches:
                    # 检查是否有匹配
                    match_found = False
                    for match in platform_matches:
                        if match in item_platform or item_platform in match:
                            match_found = True
                            break
                    if not match_found:
                        continue
                
                # 3. 情感筛选
                item_emotion = item.get("emotion", "未知")
                if emotion_type!= "全部" and item_emotion != emotion_type:
                    continue
                
                # 构建结果
               
                filtered_data.append(item)
                
                # 更新统计
                total_data_count += 1
                if item_emotion == "积极":
                    positive_count += 1
                elif item_emotion == "消极":
                    negative_count += 1
        except Exception as e:
            print(f"处理文件 {file} 时出错: {str(e)}")
    
    print(f"筛选完成: 找到{total_data_count}条符合条件的数据")
    
    return {
        "total_data_count": total_data_count,
        "today_new_count": today_new_count,
        "positive_count": positive_count,
        "negative_count": negative_count,
        "filtered_data": filtered_data,
        "code": 200,
    }


@app.get("/task_status/{task_id}")
async def get_task_status(task_id: str):
    """获取特定爬取任务的状态"""
    if task_id not in crawl_results:
        return {"status": "pending", "code": 200}
    
    return {**crawl_results[task_id], "code": 200}
# 在应用启动时，为现有的方案安排任务
@app.on_event("startup")
async def startup_event():
    keyword_groups = load_keyword_groups()
    for group in keyword_groups:
        schedule_crawling_task(group)

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
